# core/calc_parser.py
"""
Универсальный парсер расчётных файлов тарифных заявок.

Читает любые формы ЕИАС (ТЭ, ГВС, ХВС, ВО, ТКО и др.) всех версий
(машинные PJ_*-формы, человекочитаемые П.*-формы, новые «Смета»-формы),
а также произвольные Excel-сметы. Поддерживает .xlsx, .xlsm и .xlsb.

Движки чтения (каскад с fallback):
    1) openpyxl   — основной для .xlsx/.xlsm
    2) calamine   — fallback для битых .xlsx (некорректный workbook.xml и т.п.)
    3) pyxlsb     — для бинарного .xlsb

Публичный API:
    parse_workbook(bytes_or_path) -> (df, meta)
    to_llm_context(df)            -> str  (полный, по листам)
    to_llm_context_compact(df)    -> str  (компактный для промпта)
"""

from __future__ import annotations
import io, re
from typing import Dict, List, Optional, Tuple, Any
import pandas as pd

# ─────────────────────────────────────────────────────────────────────────────
# Нормализованное представление книги (не зависит от движка)
# ─────────────────────────────────────────────────────────────────────────────

class _Sheet:
    """Лист как имя + список строк (каждая строка — список ячеек)."""
    __slots__ = ("title", "rows")
    def __init__(self, title: str, rows: List[List[Any]]):
        self.title = title
        self.rows = rows
    @property
    def max_row(self) -> int:
        return len(self.rows)
    @property
    def max_column(self) -> int:
        return max((len(r) for r in self.rows), default=0)

class _Book:
    """Книга как список листов + быстрый доступ по имени."""
    def __init__(self, sheets: List[_Sheet]):
        self.sheets = sheets
        self._by_name = {s.title: s for s in sheets}
    @property
    def sheetnames(self) -> List[str]:
        return [s.title for s in self.sheets]
    def __getitem__(self, name: str) -> _Sheet:
        return self._by_name[name]


# ─────────────────────────────────────────────────────────────────────────────
# Загрузка книги (каскад движков)
# ─────────────────────────────────────────────────────────────────────────────

def _src_is_xlsb(source) -> bool:
    if isinstance(source, str):
        return source.lower().endswith(".xlsb")
    if isinstance(source, (bytes, bytearray)):
        # И .xlsx, и .xlsb — это ZIP (PK\x03\x04). Признак .xlsb —
        # наличие записи 'workbook.bin' внутри архива.
        try:
            import zipfile
            with zipfile.ZipFile(io.BytesIO(bytes(source))) as z:
                names = z.namelist()
                return any(n.endswith("workbook.bin") for n in names)
        except Exception:
            return False
    return False


def _repair_xlsx_bytes(raw: bytes) -> Optional[bytes]:
    """
    Чинит «битый» xlsx, сгенерированный кривым экспортером ЕИАС:
    дробные значения в целочисленных атрибутах workbook.xml
    (tabRatio="974.74…", firstSheet, activeTab и т.п.), на которых
    падает openpyxl с TypeError: expected <class 'int'>.
    Возвращает исправленные байты или None, если ремонт невозможен.
    """
    import zipfile
    try:
        zin = zipfile.ZipFile(io.BytesIO(raw))
    except Exception:
        return None
    _INT_ATTRS = ("tabRatio", "firstSheet", "activeTab", "windowWidth",
                  "windowHeight", "xWindow", "yWindow", "tabId",
                  "firstVisibleTab", "activeRow", "activeCol")
    try:
        buf = io.BytesIO()
        zout = zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED)
        changed = False
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith("workbook.xml"):
                txt = data.decode("utf-8", "replace")
                for attr in _INT_ATTRS:
                    new = re.sub(
                        attr + r'="(-?\d+)\.\d+"',
                        lambda m, a=attr: f'{a}="{m.group(1)}"', txt)
                    if new != txt:
                        changed = True
                        txt = new
                data = txt.encode("utf-8")
            zout.writestr(item, data)
        zout.close()
        if not changed:
            return None
        buf.seek(0)
        return buf.getvalue()
    except Exception:
        return None


def _rows_from_openpyxl(source, _errors: Optional[list] = None) -> Optional[_Book]:
    try:
        import openpyxl
    except Exception as e:
        if _errors is not None:
            _errors.append(f"openpyxl: библиотека не установлена ({e})")
        return None

    wb = None
    try:
        if isinstance(source, (bytes, bytearray)):
            wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True, read_only=True)
        else:
            wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    except Exception as e:
        # попытка авторемонта битого workbook.xml
        try:
            raw = bytes(source) if isinstance(source, (bytes, bytearray)) \
                  else open(source, "rb").read()
            fixed = _repair_xlsx_bytes(raw)
            if fixed is not None:
                wb = openpyxl.load_workbook(io.BytesIO(fixed),
                                            data_only=True, read_only=True)
        except Exception:
            wb = None
        if wb is None:
            if _errors is not None:
                _errors.append(f"openpyxl: {type(e).__name__}: {str(e)[:120]}")
            return None

    sheets = []
    try:
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            sheets.append(_Sheet(sn, rows))
    except Exception as e:
        if _errors is not None:
            _errors.append(f"openpyxl(чтение): {type(e).__name__}: {str(e)[:120]}")
        try:
            wb.close()
        except Exception:
            pass
        return None
    try:
        wb.close()
    except Exception:
        pass
    return _Book(sheets)


def _rows_from_calamine(source, _errors: Optional[list] = None) -> Optional[_Book]:
    try:
        from python_calamine import CalamineWorkbook
    except Exception as e:
        if _errors is not None:
            _errors.append(f"calamine: библиотека не установлена ({e}). "
                           f"Установите: pip install python-calamine")
        return None
    try:
        if isinstance(source, (bytes, bytearray)):
            wb = CalamineWorkbook.from_filelike(io.BytesIO(source))
        else:
            wb = CalamineWorkbook.from_path(source)
    except Exception as e:
        if _errors is not None:
            _errors.append(f"calamine: {type(e).__name__}: {str(e)[:120]}")
        return None
    sheets = []
    try:
        for sn in wb.sheet_names:
            data = wb.get_sheet_by_name(sn).to_python(skip_empty_area=False)
            rows = [list(r) for r in data]
            sheets.append(_Sheet(sn, rows))
    except Exception as e:
        if _errors is not None:
            _errors.append(f"calamine(чтение): {type(e).__name__}: {str(e)[:120]}")
        return None
    return _Book(sheets)


def _rows_from_pyxlsb(source, _errors: Optional[list] = None) -> Optional[_Book]:
    try:
        from pyxlsb import open_workbook
    except Exception as e:
        if _errors is not None:
            _errors.append(f"pyxlsb: библиотека не установлена ({e})")
        return None
    try:
        if isinstance(source, (bytes, bytearray)):
            handle = open_workbook(io.BytesIO(source))
        else:
            handle = open_workbook(source)
    except Exception as e:
        if _errors is not None:
            _errors.append(f"pyxlsb: {type(e).__name__}: {str(e)[:120]}")
        return None
    sheets = []
    try:
        with handle as wb:
            for sn in wb.sheets:
                with wb.get_sheet(sn) as sh:
                    rows = []
                    for row in sh.rows():
                        # row — список Cell(r,c,v); восстанавливаем плотный список
                        if not row:
                            rows.append([])
                            continue
                        width = max(c.c for c in row) + 1
                        line = [None] * width
                        for c in row:
                            line[c.c] = c.v
                        rows.append(line)
                    sheets.append(_Sheet(sn, rows))
    except Exception as e:
        if _errors is not None:
            _errors.append(f"pyxlsb(чтение): {type(e).__name__}: {str(e)[:120]}")
        return None
    return _Book(sheets)


def _load_book(source) -> _Book:
    """Загружает книгу, пробуя движки по очереди. При полном провале
    сообщает причину по каждому движку (в т.ч. отсутствие библиотеки)."""
    if _src_is_xlsb(source):
        order = [_rows_from_pyxlsb, _rows_from_calamine, _rows_from_openpyxl]
    else:
        order = [_rows_from_openpyxl, _rows_from_calamine, _rows_from_pyxlsb]
    errors: list = []
    for loader in order:
        book = loader(source, errors)
        if book is not None and book.sheets:
            return book
    detail = "; ".join(errors) if errors else "причина неизвестна"
    raise ValueError(
        "Не удалось прочитать файл ни одним движком (openpyxl/calamine/pyxlsb). "
        f"Детали: {detail}. "
        "Если среди причин есть «calamine не установлена» — установите "
        "python-calamine (он читает «битые» xlsx от ЕИАС): "
        "pip install python-calamine"
    )


# ─────────────────────────────────────────────────────────────────────────────
# Константы
# ─────────────────────────────────────────────────────────────────────────────

_SKIP_SHEETS = {
    "Инструкция", "Информация JSON", "Список листов", "Общие сведения",
    "Заявление", "Список территорий", "Список объектов", "Сценарии",
    "Сценарии (МСА)", "Расчет УЕ", "Транспорт", "Комментарии",
    "TEHSHEET", "TECHSHEET", "REESTR_OBJECT", "REESTR_OBJ_MO",
    "REESTR_OBJ_TRANSP", "REESTR_MO", "REESTR_ORG", "REESTR_DOP",
    "REESTR_SEP_DIV", "REESTR_IP_KS", "DICTIONARIES", "Настройка",
    "Титульный", "Документы", "Выбор метода", "Список периодов",
    "Виды деятельности", "Базы распределения", "Земельные участки",
    "Реестр жилых домов", "Реестр юр лиц", "Реестр СН", "Реестр ТСО",
    "Реестр договоров", "Информ-я о регулируемой орг-и",
    "Объекты теплоснабжения", "Список объектов теплоснабжения",
}
_SKIP_PREFIXES = ("Черновик", "REESTR_", "DICT")

# Служебные ключи ЕИАС (машинные), не являющиеся статьями
_SERVICE_KEYS = {
    "PJ_YEAR", "PJ_PF", "PJ_DOP", "PJ_PERIOD", "PJ_DOP_FIN", "PJ_DYN",
    "PJ_DYN_V", "PJ_NAME", "PJ_NAME_UNIT", "PJ_NAME_ED_IZM", "PJ_NAME_DOP",
    "PJ_NAME_FUEL", "PJ_UNIT", "dyn_names", "dyn_name", "uni_prd_data",
    "uni_org_reg", "uni_pf", "V_POK", "D_YEAR", "obj_mo", "obj_obj",
}

# Заголовки колонок — не статьи
_HEADER_STRINGS = {
    "PJ_NAME", "Наименование", "Наименование показателя", "Наименование показател",
    "№ п/п", "Ед. изм.", "Единица измерения", "Показатель", "Код",
    "всего", "горячая вода", "пар", "Не определено", "Добавить строку",
    "Комментарии и примечания", "Комментарии и примечан",
}

# Нормализация «План / факт»
_PF_NORM = {
    "предложение организации": "Предложение",
    "предложения организации": "Предложение",
    "план регулируемой организации": "Предложение",
    "план регулируемой орга": "Предложение",
    "план организации": "Предложение",
    "версия организации": "Предложение",
    "принято органом регулирования": "Принято",
    "план органа регулирования": "Принято",
    "план органа регулирова": "Принято",
    "версия регулятора": "Принято",
    "утверждено": "Принято",
    "факт по данным организации": "Факт",
    "факт, принятый органом регулирования": "Факт",
    "факт тсо": "Факт",
    "факт": "Факт",
    "ожидаемое за период по данным организации": "Ожидаемое",
    "ожидаемое": "Ожидаемое",
    "план": "План",
}

# Единицы измерения
_UNIT_PATTERN = re.compile(
    r"тыс\.?\s*руб|руб\.?\s*/?|млн\.?\s*руб|Гкал|куб\.?\s*м|кВт|тыс\.?\s*кВт|"
    r"тонн|тн\b|чел\.?|ед\.?|%|шт\.?|км|г?кал|т\.у\.т", re.I
)
_UNIT_EXACT = re.compile(
    r"^\s*(тыс\.?\s*руб\.?|руб\.?|млн\.?\s*руб\.?|Гкал.*|куб\.?\s*м.*|"
    r"тыс\.?\s*куб\.?\s*м.*|кВт.*|тыс\.?\s*кВт.*|тонн|тн|чел\.?|ед\.?|%|шт\.?|"
    r"км|руб\.?\s*/.*|т\.у\.т\.?|°?С)\s*$", re.I
)

_YEAR_MIN, _YEAR_MAX = 2010, 2045


# ─────────────────────────────────────────────────────────────────────────────
# Утилиты
# ─────────────────────────────────────────────────────────────────────────────

def _is_year(v) -> bool:
    try:
        if isinstance(v, bool):
            return False
        s = str(v).strip()
        if not re.match(r"^\d{4}(\.0)?$", s):
            return False
        y = int(float(s))
        return _YEAR_MIN <= y <= _YEAR_MAX
    except Exception:
        return False

def _norm_year(v) -> str:
    return str(int(float(str(v).strip())))

def _norm_pf(v) -> str:
    if v is None:
        return ""
    s = str(v).strip().lower()
    if not s:
        return ""
    for key, norm in _PF_NORM.items():
        if key in s:
            return norm
    return str(v).strip()[:30]

def _to_number(v) -> Optional[float]:
    """Числа и числовые строки → float. Иначе None."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace("\xa0", "").replace(" ", "")
        if not s:
            return None
        # запятая как десятичный разделитель
        if re.match(r"^-?\d+(?:[.,]\d+)?$", s):
            try:
                return float(s.replace(",", "."))
            except Exception:
                return None
    return None

def _is_unit(v) -> bool:
    return isinstance(v, str) and bool(_UNIT_EXACT.match(v.strip()))

def _is_article_name(v) -> bool:
    if not isinstance(v, str):
        return False
    s = v.strip()
    if len(s) < 3:
        return False
    if s in _SERVICE_KEYS or s in _HEADER_STRINGS:
        return False
    if _is_unit(s):
        return False
    # чистые числа / номера пунктов «1.2.3»
    if re.match(r"^[-+]?\d+([.,]\d+)*$", s):
        return False
    # коды строк ЕИАС: M4202, D0041, P_..., et_...
    if re.match(r"^[MmDdPp]\d{3,}", s):
        return False
    if re.match(r"^(check|ИТОГО_|et_|DYN_|P_|Indicators_)[A-Za-zА-Яа-я0-9_.]+$", s):
        return False
    if "::" in s or s.startswith("{") or "funcDyn" in s:
        return False
    # «true»/«false»/«да»/«нет» — флаги
    if s.lower() in {"true", "false", "да", "нет", "x", "×", "-", "—"}:
        return False
    # инвентарные/артикульные коды: длинный «хвост» из букв+цифр без пробелов
    # (напр. «Bobcat-S770H_A3P513812», «КАМАЗ-5320_X9_1234»)
    if " " not in s and re.search(r"[A-Za-z].*\d.*\d|\d{4,}", s) \
       and re.search(r"[_\-][A-Za-z0-9]{4,}", s):
        return False
    # должна быть хотя бы одна буква
    if not re.search(r"[A-Za-zА-Яа-яЁё]", s):
        return False
    return True


# ─────────────────────────────────────────────────────────────────────────────
# Детектор формата
# ─────────────────────────────────────────────────────────────────────────────

def _detect_eias(book: _Book) -> bool:
    names = set(book.sheetnames)
    if names & {"TECHSHEET", "TEHSHEET"}:
        return True
    # маркеры в ячейках
    for s in book.sheets[:60]:
        for row in s.rows[:20]:
            if not row:
                continue
            for v in row[:6]:
                if isinstance(v, str):
                    vs = v.strip()
                    if vs in ("PJ_YEAR", "PJ_PF") or vs.startswith("Indicators_") \
                       or vs.startswith("D0041") or vs.startswith("Indicators"):
                        return True
    return False


# ─────────────────────────────────────────────────────────────────────────────
# Универсальный парсер листа (работает и для ЕИАС, и для произвольных)
# ─────────────────────────────────────────────────────────────────────────────

def _scan_year_header(rows: List[List[Any]]) -> Optional[Tuple[int, Dict[int, str]]]:
    """
    Ищет строку-шапку с годами. Возвращает (индекс_строки, {col: year}).
    Берёт строку с максимальным числом year-ячеек (>=2).
    """
    best_idx, best_map = None, {}
    for i, row in enumerate(rows):
        if not row:
            continue
        ymap = {c: _norm_year(v) for c, v in enumerate(row) if _is_year(v)}
        if len(ymap) >= 2 and len(ymap) > len(best_map):
            best_idx, best_map = i, ymap
        # PJ_YEAR-форма: первая ячейка == PJ_YEAR, годы по строке
        if row and isinstance(row[0], str) and row[0].strip() == "PJ_YEAR":
            ymap2 = {c: _norm_year(v) for c, v in enumerate(row) if _is_year(v)}
            if len(ymap2) >= 1:
                return i, ymap2
    if best_idx is None:
        return None
    return best_idx, best_map


def _scan_pf_row(rows: List[List[Any]], year_idx: int, year_cols: Dict[int, str]
                 ) -> Dict[int, str]:
    """
    Ищет «План/факт» для каждой year-колонки.
    Сначала PJ_PF-строка; иначе строки вокруг year_idx с pf-метками.
    """
    # 1) PJ_PF
    for row in rows:
        if row and isinstance(row[0], str) and row[0].strip() == "PJ_PF":
            return {c: _norm_pf(row[c]) for c in year_cols if c < len(row)}
    # 2) поиск pf-меток в окне ±6 строк от year-шапки
    pf_map: Dict[int, str] = {}
    lo = max(0, year_idx - 6)
    hi = min(len(rows), year_idx + 4)
    for c in year_cols:
        for r in range(year_idx, lo - 1, -1):  # вверх от шапки
            if r >= len(rows) or c >= len(rows[r]):
                continue
            cand = rows[r][c]
            if isinstance(cand, str):
                p = _norm_pf(cand)
                if p in ("Факт", "Принято", "Предложение", "План", "Ожидаемое"):
                    pf_map[c] = p
                    break
        if c not in pf_map:
            for r in range(year_idx, hi):
                if r >= len(rows) or c >= len(rows[r]):
                    continue
                cand = rows[r][c]
                if isinstance(cand, str):
                    p = _norm_pf(cand)
                    if p in ("Факт", "Принято", "Предложение", "План", "Ожидаемое"):
                        pf_map[c] = p
                        break
    return pf_map


_NAME_HEADER = re.compile(r"^\s*наименование", re.I)
_COMMENT_HEADER = re.compile(r"коммент|примечан|обоснован", re.I)

def _find_header_name_col(rows: List[List[Any]], year_idx: int,
                          year_cols: Dict[int, str]) -> Optional[int]:
    """
    Прямой поиск колонки по заголовку «Наименование …».
    Ищет в окне строк около year-шапки. Возвращает None, если не найдено.
    """
    lo = max(0, year_idx - 8)
    hi = min(len(rows), year_idx + 4)
    for r in range(lo, hi):
        row = rows[r]
        if not row:
            continue
        for c, v in enumerate(row):
            if c in year_cols or not isinstance(v, str):
                continue
            if _NAME_HEADER.match(v.strip()):
                return c
    return None


def _comment_cols(rows: List[List[Any]], year_idx: int) -> set:
    """Колонки-комментарии (по заголовку) — их нельзя брать как имя статьи."""
    cols = set()
    lo = max(0, year_idx - 8)
    hi = min(len(rows), year_idx + 4)
    for r in range(lo, hi):
        row = rows[r]
        if not row:
            continue
        for c, v in enumerate(row):
            if isinstance(v, str) and _COMMENT_HEADER.search(v):
                cols.add(c)
    return cols


def _find_name_col(rows: List[List[Any]], data_start: int,
                   year_cols: Dict[int, str], ban: set = frozenset()
                   ) -> Optional[int]:
    """
    Колонка с наименованиями статей.
    Берёт колонку с наибольшим числом коротких/средних article-имён
    (длина 8..120), исключая колонки-комментарии (паб. ban).
    Длинные «простыни» (>120 симв.) штрафуются — это обоснования, не статьи.
    """
    counts: Dict[int, int] = {}
    longpen: Dict[int, int] = {}
    for row in rows[data_start: data_start + 200]:
        if not row:
            continue
        for c, v in enumerate(row):
            if c in year_cols or c in ban or not isinstance(v, str):
                continue
            s = v.strip()
            if not _is_article_name(s):
                continue
            L = len(s)
            if L < 8:
                continue
            if L > 120:
                longpen[c] = longpen.get(c, 0) + 1
                continue
            counts[c] = counts.get(c, 0) + 1
    if not counts:
        return None
    # выбираем колонку с макс. числом «нормальных» имён;
    # при равенстве — с меньшим числом «простыней»
    def key(c):
        return (counts[c], -longpen.get(c, 0))
    return max(counts, key=key)


def _find_unit_col(rows: List[List[Any]], data_start: int,
                   name_col: int, year_cols: Dict[int, str]) -> Optional[int]:
    """Колонка с единицами измерения."""
    scores: Dict[int, int] = {}
    for row in rows[data_start: data_start + 120]:
        if not row:
            continue
        for c, v in enumerate(row):
            if c == name_col or c in year_cols:
                continue
            if _is_unit(v):
                scores[c] = scores.get(c, 0) + 1
    if not scores:
        return None
    return max(scores, key=scores.__getitem__)


def _parse_sheet(sheet: _Sheet) -> List[Dict]:
    rows = sheet.rows
    if not rows:
        return []

    hdr = _scan_year_header(rows)
    if hdr is None:
        return []  # без годов лист не разбираем (избегаем мусора)
    year_idx, year_cols = hdr
    pf_map = _scan_pf_row(rows, year_idx, year_cols)

    data_start = year_idx + 1
    ban = _comment_cols(rows, year_idx)
    # 1) прямой поиск по заголовку «Наименование»
    name_col = _find_header_name_col(rows, year_idx, year_cols)
    # 2) эвристика, если заголовка нет
    if name_col is None:
        name_col = _find_name_col(rows, data_start, year_cols, ban)
    if name_col is None:
        name_col = _find_name_col(rows, 0, year_cols, ban)
    if name_col is None:
        return []
    unit_col = _find_unit_col(rows, data_start, name_col, year_cols)

    records: List[Dict] = []
    title = sheet.title

    for row in rows[data_start:]:
        if not row:
            continue
        # имя статьи — строго из name_col
        if name_col >= len(row):
            continue
        cell = row[name_col]
        if not _is_article_name(cell):
            continue
        article = str(cell).strip()
        # отсекаем «простыни»-обоснования (не статьи)
        if len(article) > 160:
            continue

        # единица
        unit = ""
        if unit_col is not None and unit_col < len(row):
            if _is_unit(row[unit_col]):
                unit = str(row[unit_col]).strip()

        # значения
        for c, year in year_cols.items():
            if c >= len(row):
                continue
            num = _to_number(row[c])
            if num is None or num == 0:
                continue
            records.append({
                "sheet":   title,
                "article": article,
                "period":  year,
                "pf":      pf_map.get(c, ""),
                "value":   num,
                "unit":    unit,
            })
    return records


# ─────────────────────────────────────────────────────────────────────────────
# Произвольные формы (УРТ, сметы РСО): год/«план-факт» зашиты в ТЕКСТЕ шапки
# колонки, напр. «Утверждено на 2016 год», «Факт 2016 года»,
# «Предложение РСО на 2018». Иногда разбито на 2 строки шапки:
# «Утверждено» + «на 2016 г.».  Извлекаем (год, pf) из склеенного текста.
# ─────────────────────────────────────────────────────────────────────────────

_YEAR_IN_TEXT = re.compile(r"\b(19[9]\d|20[0-4]\d)\b")
_YEAR_RANGE = re.compile(r"\b(20[0-4]\d)\s*[-–—]\s*(20[0-4]\d)\b")
# мусорные шапки: ссылки на НПА, даты приказов — год там не относится к данным
_DECREE_NOISE = re.compile(
    r"приказ|пост[ае]новл|ф[сc]т\b|фа[сc]\b|методич|№\s*\d|"
    r"\d{1,2}\.\d{1,2}\.\d{4}", re.I)

def _year_from_header(text: str) -> str:
    """Извлекает год периода из текста шапки колонки. '' если не относится к данным."""
    if not text:
        return ""
    if _DECREE_NOISE.search(text):
        # вырезаем хвост ссылки на НПА (от ключевого слова и далее),
        # а также даты/номера приказов в любом месте строки
        text = re.sub(r"(приказ|пост[ае]новл|методич|норматив\w*\s+в\s+соответ)"
                      r".{0,60}", " ", text, flags=re.I)
        text = re.sub(r"\b\d{1,2}\.\d{1,2}\.\d{4}\b", " ", text)
        text = re.sub(r"№\s*\d+[-\w/]*", " ", text)
        text = re.sub(r"\bот\s+\d{2,4}\b", " ", text, flags=re.I)
    rng = _YEAR_RANGE.search(text)
    if rng:
        return rng.group(2)  # год периода регулирования = конец диапазона
    m = _YEAR_IN_TEXT.search(text)
    return m.group(1) if m else ""

# pf-маркеры в тексте шапки (порядок важен — длинные раньше)
_PF_TEXT = [
    (re.compile(r"принят|утвержд|урт|орган(?:а|ом)?\s+регул", re.I), "Принято"),
    (re.compile(r"предложени|план\s+рсо|рсо\b|заявк", re.I),        "Предложение"),
    (re.compile(r"ожидаем", re.I),                                  "Ожидаемое"),
    (re.compile(r"\bфакт", re.I),                                   "Факт"),
    (re.compile(r"\bплан", re.I),                                   "План"),
]

def _pf_from_text(text: str) -> str:
    for rx, name in _PF_TEXT:
        if rx.search(text):
            return name
    return ""


def _build_column_headers(rows: List[List[Any]], header_top: int, header_bot: int,
                          name_col: int) -> Dict[int, Tuple[str, str]]:
    """
    Склеивает текст шапки по каждой колонке в строках [header_top, header_bot)
    и извлекает (year, pf). Числовые ячейки-годы тоже учитываются как часть
    шапки. Возвращает {col: (year, pf)} для колонок с найденным годом или pf.
    """
    width = max((len(r) for r in rows[header_top:header_bot]), default=0)
    combined: Dict[int, str] = {}
    for r in range(header_top, header_bot):
        if r >= len(rows):
            break
        row = rows[r]
        for c in range(width):
            if c == name_col:
                continue
            v = row[c] if c < len(row) else None
            piece = None
            if isinstance(v, str) and v.strip():
                piece = v.strip()
            elif _is_year(v):                       # числовой год в шапке
                piece = _norm_year(v)
            if piece:
                combined[c] = (combined.get(c, "") + " " + piece).strip()

    headers: Dict[int, Tuple[str, str]] = {}
    for c, text in combined.items():
        year = _year_from_header(text)
        pf = _pf_from_text(text)
        if year or pf:
            headers[c] = (year, pf)
    return headers


def _fill_missing_years(headers: Dict[int, Tuple[str, str]]
                        ) -> Dict[int, Tuple[str, str]]:
    """
    Если у части колонок год не распознан (напр. шапка «(РСО) / УРТ»),
    но есть соседи с годом — протягиваем ближайший известный год слева.
    """
    if not headers:
        return headers
    cols = sorted(headers)
    last_year = ""
    out = dict(headers)
    for c in cols:
        y, pf = out[c]
        if y:
            last_year = y
        elif last_year:
            out[c] = (last_year, pf)
    return out


def _parse_columnar_sheet(sheet: _Sheet) -> List[Dict]:
    """
    Парсер произвольных форм с годом/pf в тексте (или числом) шапки колонки.
    """
    rows = sheet.rows
    if not rows:
        return []
    title = sheet.title

    # 1) колонка наименований — по заголовку, иначе текстовая эвристика
    name_col = None
    name_hdr = re.compile(r"наименован|показател|параметр|статья", re.I)
    for r in range(min(15, len(rows))):
        row = rows[r]
        if not row:
            continue
        for c, v in enumerate(row):
            if isinstance(v, str) and name_hdr.search(v):
                name_col = c
                break
        if name_col is not None:
            break
    if name_col is None:
        name_col = _find_name_col(rows, 0, {}, frozenset())
    if name_col is None:
        return []

    # 2) начало данных: строка с именем статьи в name_col + числа справа
    data_start = None
    for r in range(min(40, len(rows))):
        row = rows[r]
        if not row or name_col >= len(row):
            continue
        if _is_article_name(row[name_col]):
            nums = [c for c, v in enumerate(row)
                    if c != name_col and _to_number(v) not in (None, 0)]
            if nums:
                data_start = r
                break
    if data_start is None:
        return []

    # 3) шапка — узкое окно над данными (до 8 строк), чтобы не цеплять
    #    далёкие титульные строки со ссылками на приказы
    header_top = max(0, data_start - 8)
    headers = _build_column_headers(rows, header_top, data_start, name_col)
    headers = _fill_missing_years(headers)
    data_cols = set(headers)

    unit_col = _find_unit_col(rows, data_start, name_col, data_cols)

    records: List[Dict] = []
    for row in rows[data_start:]:
        if not row or name_col >= len(row):
            continue
        if not _is_article_name(row[name_col]):
            continue
        article = str(row[name_col]).strip()
        if len(article) > 160:
            continue

        unit = ""
        if unit_col is not None and unit_col < len(row) and _is_unit(row[unit_col]):
            unit = str(row[unit_col]).strip()

        if data_cols:
            # строго по распознанным колонкам данных — год привязан верно
            for c, (year, pf) in headers.items():
                if c >= len(row):
                    continue
                num = _to_number(row[c])
                if num is None or num == 0:
                    continue
                records.append({
                    "sheet": title, "article": article,
                    "period": year, "pf": pf, "value": num, "unit": unit,
                })
        else:
            # год в шапке не найден вообще → период неизвестен (помечаем «н/д»)
            for c, v in enumerate(row):
                if c == name_col or c == unit_col:
                    continue
                num = _to_number(v)
                if num is None or num == 0:
                    continue
                records.append({
                    "sheet": title, "article": article,
                    "period": "н/д", "pf": "", "value": num, "unit": unit,
                })
    return records


# обратная совместимость имени
def _parse_arbitrary_sheet(sheet: _Sheet) -> List[Dict]:
    return _parse_columnar_sheet(sheet)


# ─────────────────────────────────────────────────────────────────────────────
# Обход книги
# ─────────────────────────────────────────────────────────────────────────────

def _skip_sheet(name: str) -> bool:
    if name in _SKIP_SHEETS:
        return True
    return any(name.startswith(p) for p in _SKIP_PREFIXES)


def _parse_book(book: _Book, is_eias: bool) -> Tuple[pd.DataFrame, Dict]:
    all_records: List[Dict] = []
    parsed = skipped = 0
    for sheet in book.sheets:
        if _skip_sheet(sheet.title):
            skipped += 1
            continue
        if sheet.max_row < 3 or sheet.max_column < 2:
            skipped += 1
            continue
        if is_eias:
            recs = _parse_sheet(sheet)
        else:
            # произвольные формы: год/pf в тексте шапки колонки
            recs = _parse_columnar_sheet(sheet)
            if not recs:
                # вдруг это лист с обычной year-шапкой (числовой год в ячейке)
                recs = _parse_sheet(sheet)
        if recs:
            all_records.extend(recs)
            parsed += 1
        else:
            skipped += 1

    df = pd.DataFrame(all_records) if all_records else pd.DataFrame(
        columns=["sheet", "article", "period", "pf", "value", "unit"])
    meta = {
        "format":         "ЕИАС" if is_eias else "Произвольный",
        "sheets_parsed":  parsed,
        "sheets_skipped": skipped,
        "records_total":  len(df),
    }
    return df, meta


# ─────────────────────────────────────────────────────────────────────────────
# Постобработка
# ─────────────────────────────────────────────────────────────────────────────

def _deduplicate(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    pf_priority = {"Принято": 0, "Предложение": 1, "Факт": 2,
                   "План": 3, "Ожидаемое": 4, "": 5}
    df = df.copy()
    df["_rank"] = df["pf"].map(lambda x: pf_priority.get(x, 6))
    df = (df.sort_values("_rank")
            .drop_duplicates(subset=["sheet", "article", "period", "pf"], keep="first")
            .drop(columns=["_rank"]))
    return df.reset_index(drop=True)


def _clean(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    df["article"] = df["article"].astype(str).str.strip().str[:200]
    df = df[df["article"].str.len() >= 3]
    df = df[~df["article"].str.startswith("{")]
    df = df[df["article"].str.contains(r"[A-Za-zА-Яа-яЁё]", regex=True)]
    return df.reset_index(drop=True)


# ─────────────────────────────────────────────────────────────────────────────
# Публичный API
# ─────────────────────────────────────────────────────────────────────────────

def parse_workbook(source) -> Tuple[pd.DataFrame, Dict]:
    """
    Главная точка входа.
    source: bytes | str (путь). Возвращает (df, meta) — только ненулевые статьи.
    """
    book = _load_book(source)
    is_eias = _detect_eias(book)
    df, meta = _parse_book(book, is_eias)

    if not df.empty:
        df = _clean(df)
        df = _deduplicate(df)
        df = df[df["value"] != 0].reset_index(drop=True)

    meta["records_final"]  = len(df)
    meta["articles_final"] = int(df["article"].nunique()) if not df.empty else 0
    return df, meta


def to_llm_context(df: pd.DataFrame, max_articles: int = 0) -> str:
    """Полный формат для LLM — сгруппирован по листу и статье."""
    if df.empty:
        return "Расчётный файл не содержит данных."
    lines = []
    for sheet, sdf in df.groupby("sheet", sort=False):
        lines.append(f"\n# {sheet}")
        ranked = sdf.groupby("article")["value"].max().abs().sort_values(ascending=False)
        top = ranked.head(max_articles).index if max_articles > 0 else ranked.index
        for article in top:
            adf = sdf[sdf["article"] == article].sort_values("period")
            lines.append(f"\n★ {article}")
            for _, row in adf.iterrows():
                period = row["period"] or "—"
                pf = f" ({row['pf']})" if row["pf"] else ""
                val = row["value"]
                unit = row["unit"] or "тыс.руб."
                if abs(val) >= 1_000_000:
                    vs = f"{val/1_000_000:.2f} млн"
                elif abs(val) >= 1_000:
                    vs = f"{val:,.0f}"
                else:
                    vs = f"{val:.2f}"
                lines.append(f"  {period}{pf}: {vs} {unit}")
    return "\n".join(lines)


def to_llm_context_compact(df: pd.DataFrame, max_articles: int = 0) -> str:
    """Компактный формат для промпта."""
    if df.empty:
        return "Нет данных."
    ranked = df.groupby("article")["value"].max().abs().sort_values(ascending=False)
    top = ranked.head(max_articles).index if max_articles > 0 else ranked.index
    lines = []
    for article in top:
        adf = df[df["article"] == article].sort_values("period")
        parts = []
        for _, row in adf.iterrows():
            period = row["period"] or "—"
            pf = f"/{row['pf'][:3]}" if row["pf"] else ""
            val = row["value"]
            vs = f"{val/1_000_000:.1f}млн" if abs(val) >= 1_000_000 else f"{val:,.0f}"
            parts.append(f"{period}{pf}={vs}")
        lines.append(f"  {article}: {', '.join(parts)}")
    return "\n".join(lines)